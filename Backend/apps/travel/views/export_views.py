"""
Travel Application Export View
Generates an Excel file of travel applications filtered by trip start date range.
Accessible by Admin users only.
"""

import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from apps.authentication.permissions import IsAdminUser
from apps.travel.models.application import TravelApplication, TripDetails
from utils.response_formatter import error_response

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(value: str, param_name: str):
    """Parse a YYYY-MM-DD string; raise ValueError with a clear message on failure."""
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        raise ValueError(f"'{param_name}' must be a valid date in YYYY-MM-DD format.")


def _format_time(t):
    """Return HH:MM string or empty string for None."""
    if t is None:
        return ""
    return t.strftime("%H:%M")


def _format_date(d):
    """Return DD-Mon-YYYY string or empty string for None."""
    if d is None:
        return ""
    return d.strftime("%d-%b-%Y")


def _get_travel_request_id(app: TravelApplication) -> str:
    return f"TR/TSF/{app.created_at.year}/{app.id:07d}"


def _build_row(app: TravelApplication, first_trip, last_trip) -> list:
    """Build a single data row for the Excel sheet."""
    employee = app.employee
    full_name = f"{employee.first_name} {employee.last_name}".strip() or employee.username

    origin = (
        first_trip.from_location.city_name
        if first_trip and first_trip.from_location
        else ""
    )
    destination = (
        last_trip.to_location.city_name
        if last_trip and last_trip.to_location
        else ""
    )

    return [
        _get_travel_request_id(app),
        full_name,
        app.purpose or "",
        app.get_status_display(),
        origin,
        destination,
        _format_date(first_trip.departure_date if first_trip else None),
        _format_time(first_trip.start_time if first_trip else None),
        _format_date(last_trip.return_date if last_trip else None),
        _format_time(last_trip.end_time if last_trip else None),
    ]


def _build_preview_item(app: TravelApplication, first_trip, last_trip) -> dict:
    """Build a lightweight dict for the JSON preview response."""
    employee = app.employee
    full_name = f"{employee.first_name} {employee.last_name}".strip() or employee.username

    return {
        "travel_request_id": _get_travel_request_id(app),
        "employee_name": full_name,
        "purpose": app.purpose or "",
        "status": app.get_status_display(),
        "origin": (
            first_trip.from_location.city_name
            if first_trip and first_trip.from_location
            else ""
        ),
        "destination": (
            last_trip.to_location.city_name
            if last_trip and last_trip.to_location
            else ""
        ),
        "trip_start_date": _format_date(first_trip.departure_date if first_trip else None),
        "trip_start_time": _format_time(first_trip.start_time if first_trip else None),
        "trip_end_date": _format_date(last_trip.return_date if last_trip else None),
        "trip_end_time": _format_time(last_trip.end_time if last_trip else None),
    }


def _fetch_applications(start_date, end_date):
    """
    Return a list of tuples:
        (TravelApplication, first_trip_or_None, last_trip_or_None)

    Only includes applications whose EARLIEST trip departure_date (trip start date)
    falls within [start_date, end_date].  Later legs of a multi-leg trip that happen
    to fall in the range do NOT qualify the application on their own.
    """
    from django.db.models import Min

    # Annotate each application with its earliest departure_date across all legs,
    # then keep only those whose minimum departure_date is within the range.
    app_ids = (
        TripDetails.objects
        .values("travel_application_id")
        .annotate(earliest_start=Min("departure_date"))
        .filter(earliest_start__range=(start_date, end_date))
        .values_list("travel_application_id", flat=True)
    )

    applications = (
        TravelApplication.objects
        .filter(id__in=app_ids)
        .select_related("employee")
        .prefetch_related("trip_details__from_location", "trip_details__to_location")
        .order_by("id")
    )

    result = []
    for app in applications:
        # first trip  = leg with the earliest departure_date (trip start)
        trips = list(app.trip_details.order_by("departure_date", "id"))
        first_trip = trips[0] if trips else None

        # last trip = leg with the latest return_date (trip end)
        last_trip = (
            sorted(trips, key=lambda t: (t.return_date or t.departure_date, t.id))[-1]
            if trips else None
        )
        result.append((app, first_trip, last_trip))

    return result


def _generate_excel(rows: list, start_date, end_date) -> bytes:
    """Build and return the Excel workbook as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Travel Applications"

    # ---- Styles ----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    meta_font = Font(bold=True, size=10)
    meta_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    # ---- Title / Meta rows ----
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Travel Applications Report"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    range_cell = ws["A2"]
    range_cell.value = (
        f"Date Range: {_format_date(start_date)}  →  {_format_date(end_date)}"
        f"     |     Total Records: {len(rows)}"
    )
    range_cell.font = meta_font
    range_cell.fill = meta_fill
    range_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ---- Column headers ----
    HEADERS = [
        "Travel Request ID",
        "Employee Name",
        "Purpose",
        "Application Status",
        "Trip Origin",
        "Trip Destination",
        "Trip Start Date",
        "Trip Start Time",
        "Trip End Date",
        "Trip End Time",
    ]

    COLUMN_WIDTHS = [22, 24, 40, 24, 22, 22, 18, 16, 18, 16]

    header_row = 3
    for col_idx, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 30

    # ---- Data rows ----
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 3))
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # ---- Freeze panes below header ----
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ---- Write to bytes ----
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

class TravelApplicationExportPreviewView(APIView):
    """
    GET /api/travel/admin/export/preview/?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD

    Returns a JSON preview of the records that will be exported.
    Used by the frontend to show a summary table before downloading.
    """

    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        start_date_str = request.query_params.get("start_date", "").strip()
        end_date_str = request.query_params.get("end_date", "").strip()

        # --- Validate inputs ---
        if not start_date_str or not end_date_str:
            return error_response(
                message="Both 'start_date' and 'end_date' query parameters are required.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = _parse_date(start_date_str, "start_date")
            end_date = _parse_date(end_date_str, "end_date")
        except ValueError as exc:
            return error_response(message=str(exc), status_code=status.HTTP_400_BAD_REQUEST)

        if start_date > end_date:
            return error_response(
                message="'start_date' cannot be after 'end_date'.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        try:
            data = _fetch_applications(start_date, end_date)
        except Exception as exc:
            logger.error("Error fetching travel applications for preview: %s", exc, exc_info=True)
            return error_response(
                message="An unexpected error occurred while fetching data.",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        preview_items = [_build_preview_item(app, ft, lt) for app, ft, lt in data]

        # Status breakdown summary
        status_counts: dict = {}
        for app, _, _ in data:
            label = app.get_status_display()
            status_counts[label] = status_counts.get(label, 0) + 1

        return Response(
            {
                "success": True,
                "message": "Preview data fetched successfully.",
                "data": {
                    "total": len(preview_items),
                    "start_date": str(start_date),
                    "end_date": str(end_date),
                    "status_summary": status_counts,
                    "records": preview_items,
                },
            },
            status=status.HTTP_200_OK,
        )


class TravelApplicationExportView(APIView):
    """
    GET /api/travel/admin/export/?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD

    Streams an Excel (.xlsx) file containing travel applications
    whose earliest trip departure date falls within the given range.
    """

    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        start_date_str = request.query_params.get("start_date", "").strip()
        end_date_str = request.query_params.get("end_date", "").strip()

        # --- Validate inputs ---
        if not start_date_str or not end_date_str:
            return error_response(
                message="Both 'start_date' and 'end_date' query parameters are required.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = _parse_date(start_date_str, "start_date")
            end_date = _parse_date(end_date_str, "end_date")
        except ValueError as exc:
            return error_response(message=str(exc), status_code=status.HTTP_400_BAD_REQUEST)

        if start_date > end_date:
            return error_response(
                message="'start_date' cannot be after 'end_date'.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        try:
            data = _fetch_applications(start_date, end_date)
        except Exception as exc:
            logger.error("Error fetching travel applications for export: %s", exc, exc_info=True)
            return error_response(
                message="An unexpected error occurred while fetching data.",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        rows = [_build_row(app, ft, lt) for app, ft, lt in data]

        try:
            excel_bytes = _generate_excel(rows, start_date, end_date)
        except Exception as exc:
            logger.error("Error generating Excel file: %s", exc, exc_info=True)
            return error_response(
                message="Failed to generate Excel file.",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        filename = (
            f"Travel_Applications_{start_date.strftime('%d%b%Y')}"
            f"_to_{end_date.strftime('%d%b%Y')}.xlsx"
        )

        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        return response
