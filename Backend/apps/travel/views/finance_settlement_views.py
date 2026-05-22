import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from apps.authentication.mixins import BranchFilterMixin
from apps.authentication.spoc_utils import get_user_assigned_location_ids
from apps.travel.services.settlement_overdue import (
    SETTLEMENT_OVERDUE_EXPORT_HEADERS,
    apply_settlement_overdue_filters,
    get_settlement_overdue_base_queryset,
    serialize_settlement_overdue_row,
    settlement_overdue_row_to_excel,
)
from apps.travel.views.dashboards import FINANCE_SPOC_ROLE, _verify_finance_permissions
from utils.pagination import StandardResultsSetPagination
from utils.response_formatter import error_response, success_response


class SettlementOverdueListView(BranchFilterMixin, APIView):
    """Travel completed, claim not raised, settlement period lapsed."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _verify_finance_permissions(request.user):
            return error_response(message="Permission denied", status_code=403)

        qs = get_settlement_overdue_base_queryset()
        qs = self.apply_branch_filter(
            qs, request.user, employee_field="employee", spoc_role_name=FINANCE_SPOC_ROLE
        )

        location_id = request.query_params.get("location_id")
        loc_id = None
        if location_id and location_id != "all":
            try:
                loc_id = int(location_id)
            except (TypeError, ValueError):
                return error_response(message="Invalid location_id", status_code=400)
            allowed = get_user_assigned_location_ids(
                request.user, role_name=FINANCE_SPOC_ROLE, include_base_location=True
            )
            if allowed is not None and loc_id not in allowed:
                qs = qs.none()

        qs = apply_settlement_overdue_filters(
            qs,
            location_id=loc_id,
            search=request.query_params.get("search"),
        )

        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(qs, request)
        if page is not None:
            results = [serialize_settlement_overdue_row(app) for app in page]
            total = paginator.page.paginator.count
        else:
            results = [serialize_settlement_overdue_row(app) for app in qs]
            total = len(results)

        return success_response(
            data={"count": total, "results": results},
            message="Settlement overdue applications retrieved",
        )


class SettlementOverdueExportView(BranchFilterMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _verify_finance_permissions(request.user):
            return error_response(message="Permission denied", status_code=403)

        qs = get_settlement_overdue_base_queryset()
        qs = self.apply_branch_filter(
            qs, request.user, employee_field="employee", spoc_role_name=FINANCE_SPOC_ROLE
        )

        location_id = request.query_params.get("location_id")
        loc_id = None
        if location_id and location_id != "all":
            try:
                loc_id = int(location_id)
            except (TypeError, ValueError):
                return error_response(message="Invalid location_id", status_code=400)
            allowed = get_user_assigned_location_ids(
                request.user, role_name=FINANCE_SPOC_ROLE, include_base_location=True
            )
            if allowed is not None and loc_id not in allowed:
                qs = qs.none()

        qs = apply_settlement_overdue_filters(
            qs,
            location_id=loc_id,
            search=request.query_params.get("search"),
        )

        rows = [
            settlement_overdue_row_to_excel(serialize_settlement_overdue_row(app))
            for app in qs
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settlement Overdue"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col, header in enumerate(SETTLEMENT_OVERDUE_EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 2, 14)

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"settlement_overdue_claims_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
