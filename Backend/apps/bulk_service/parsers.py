import csv
import openpyxl
from abc import ABC, abstractmethod
from typing import List, Generator, Dict, Any
from rest_framework.exceptions import ValidationError

class BaseParser(ABC):
    """
    Abstract base parser for handling different file formats.
    """
    def __init__(self, file):
        self.file = file

    @abstractmethod
    def parse(self) -> Generator[Dict[str, Any], None, None]:
        """
        Yields rows as dictionaries {header: value}.
        """
        pass

    @abstractmethod
    def get_headers(self) -> List[str]:
        """
        Returns a list of headers from the file.
        """
        pass

class CSVParser(BaseParser):
    def __init__(self, file):
        super().__init__(file)
        self.reader = None
        self._headers = []
        self._initialize_reader()

    def _initialize_reader(self):
        # Ensure file is in text mode or decode if bytes
        if hasattr(self.file, 'read'):
            if hasattr(self.file, 'seek'):
                self.file.seek(0)
            content = self.file.read()
            if isinstance(content, bytes):
                decoded = False
                for encoding in ['utf-8-sig', 'cp1252', 'latin-1']:
                    try:
                        content = content.decode(encoding)
                        decoded = True
                        break
                    except UnicodeDecodeError:
                         continue
                
                if not decoded:
                    raise ValidationError("Could not decode file. Please save as UTF-8 CSV.")

            self.file.seek(0)
            from io import StringIO
            self.file_obj = StringIO(content)
        
        try:
            self.reader = csv.DictReader(self.file_obj)
            self._headers = [h.strip() for h in self.reader.fieldnames or []] if self.reader.fieldnames else []
        except Exception as e:
            raise ValidationError(f"Invalid CSV file: {str(e)}")

    def get_headers(self) -> List[str]:
        return self._headers

    def parse(self) -> Generator[Dict[str, Any], None, None]:
        # Reset stream to ensure we start from the beginning
        if not self.file_obj:
            return
            
        self.file_obj.seek(0)
        reader = csv.DictReader(self.file_obj)
                
        count = 0
        for row in reader:
            count += 1
            # Clean up keys and values
            yield {k.strip(): v.strip() if v else None for k, v in row.items() if k}

class XLSXParser(BaseParser):
    def __init__(self, file):
        super().__init__(file)
        self.workbook = None
        self.sheet = None
        self._headers = []
        self._initialize_workbook()

    def _initialize_workbook(self):
        try:
            if hasattr(self.file, 'seek'):
                self.file.seek(0)
            self.workbook = openpyxl.load_workbook(self.file, data_only=True)
            self.sheet = self.workbook.active
            
            # Read first row as headers
            header_row = next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                self._headers = [str(h).strip() for h in header_row if h is not None]
        except Exception as e:
            raise ValidationError(f"Invalid Excel file: {str(e)}")

    def get_headers(self) -> List[str]:
        return self._headers

    def parse(self) -> Generator[Dict[str, Any], None, None]:
        if not self.sheet:
            return

        headers = self.get_headers()
        # Iterate from 2nd row
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            # Map row values to headers, handling potential missing values
            row_dict = {}
            for i, header in enumerate(headers):
                idx = i  # simple mapping since we generated headers from row 1
                val = row[idx] if idx < len(row) else None
                # Basic cleaning
                if val is not None:
                     if isinstance(val, str):
                         val = val.strip()
                row_dict[header] = val
            yield row_dict

def get_parser(file, file_extension: str) -> BaseParser:
    ext = file_extension.lower().replace('.', '')
    if ext == 'csv':
        return CSVParser(file)
    elif ext in ['xlsx', 'xls']:
        return XLSXParser(file)
    else:
        raise ValidationError(f"Unsupported file extension: {ext}")
